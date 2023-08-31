import streamlit as st
import tldextract
from openpyxl import load_workbook

def get_unique_links(existing_links, new_links):
    existing_domains = {tldextract.extract(link).domain for link in existing_links}
    unique_new_links = [link for link in new_links if tldextract.extract(link).domain not in existing_domains]
    return unique_new_links

def main():
    st.title("Vikash Goyal's New Project: Find Links")

    st.write("Upload an Excel file and provide a list of new links.")
    
    excel_file = st.file_uploader("Upload Excel file", type=["xlsx"])
    new_links_input = st.text_area("Enter new links (one link per line)")

    if excel_file and new_links_input:
        new_links = [link.strip() for link in new_links_input.split("\n") if link.strip()]
        
        existing_links = []
        if excel_file:
            workbook = load_workbook(excel_file)
            sheet = workbook.active
            existing_links = [sheet.cell(row=row_num, column=1).value for row_num in range(2, sheet.max_row + 1)]
        
        unique_links = get_unique_links(existing_links, new_links)
        
        st.write("Unique Links:")
        for link in unique_links:
            st.write(link)

if __name__ == "__main__":
    main()
